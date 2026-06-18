"""
Auth module — login + admin-only account management (SQLite/local version).

Endpoints:
  POST   /api/auth/login                     — email + password → JWT
  POST   /api/auth/admin/create-teacher      — create teacher / hod / admin account
  POST   /api/auth/admin/create-student      — create single student account
  POST   /api/auth/admin/preview-students    — parse xlsx, return rows (no DB changes)
  POST   /api/auth/admin/upload-students     — bulk-create students from xlsx
  DELETE /api/auth/admin/users/{user_id}     — delete an account
"""
from __future__ import annotations

import os
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Optional

import jwt
import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from passlib.context import CryptContext
from pydantic import BaseModel

from jwt_auth import verify_token, require_admin
from database import fetchone, execute, new_id

router = APIRouter(prefix="/api/auth", tags=["auth"])

_SECRET = os.getenv("JWT_SECRET", "")
_TOKEN_TTL_DAYS = 7

_pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")


def _hash_password(password: str) -> str:
    return _pwd_ctx.hash(password)


def _verify_password(plain: str, hashed: str) -> bool:
    return _pwd_ctx.verify(plain, hashed)


def _issue_token(user_id: str, email: str) -> str:
    if not _SECRET:
        raise HTTPException(500, "JWT_SECRET not configured on the server.")
    payload = {
        "sub": user_id,
        "email": email,
        "iat": datetime.now(timezone.utc),
        "exp": datetime.now(timezone.utc) + timedelta(days=_TOKEN_TTL_DAYS),
    }
    return jwt.encode(payload, _SECRET, algorithm="HS256")


# ── Schemas ───────────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    email: str
    password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class ResetPasswordRequest(BaseModel):
    new_password: str


class CreateTeacherRequest(BaseModel):
    name: str
    email: str
    password: str
    role: str  # 'teacher' | 'hod' | 'admin'


class CreateStudentRequest(BaseModel):
    name: str
    email: str
    password: str
    usn: Optional[str] = None


# ── Login ─────────────────────────────────────────────────────────────────────

@router.post("/login")
def login(body: LoginRequest):
    email = body.email.strip().lower()
    user = fetchone("SELECT * FROM profiles WHERE email = ?", (email,))
    if not user or not _verify_password(body.password, user.get("password_hash", "")):
        raise HTTPException(401, "Invalid email or password.")
    token = _issue_token(user["id"], email)
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {"id": user["id"], "email": email, "role": user.get("role")},
    }


# ── Change own password ───────────────────────────────────────────────────────

@router.post("/change-password")
def change_password(body: ChangePasswordRequest, user: dict = Depends(verify_token)):
    if len(body.new_password) < 6:
        raise HTTPException(400, "New password must be at least 6 characters.")
    row = fetchone("SELECT * FROM profiles WHERE id = ?", (user["sub"],))
    if not row:
        raise HTTPException(404, "User not found.")
    if not _verify_password(body.current_password, row.get("password_hash", "")):
        raise HTTPException(401, "Current password is incorrect.")
    execute(
        "UPDATE profiles SET password_hash = ? WHERE id = ?",
        (_hash_password(body.new_password), user["sub"]),
    )
    return {"message": "Password changed successfully."}


# ── Admin: reset any user's password ─────────────────────────────────────────

@router.post("/admin/users/{user_id}/reset-password")
def admin_reset_password(user_id: str, body: ResetPasswordRequest, user: dict = Depends(verify_token)):
    require_admin(user)
    if len(body.new_password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters.")
    if not fetchone("SELECT id FROM profiles WHERE id = ?", (user_id,)):
        raise HTTPException(404, "User not found.")
    execute(
        "UPDATE profiles SET password_hash = ? WHERE id = ?",
        (_hash_password(body.new_password), user_id),
    )
    return {"message": "Password reset successfully."}


# ── Admin: create teacher / hod / admin account ───────────────────────────────

@router.post("/admin/create-teacher")
def create_teacher(body: CreateTeacherRequest, user: dict = Depends(verify_token)):
    require_admin(user)

    valid_roles = {"teacher", "hod", "admin"}
    if body.role not in valid_roles:
        raise HTTPException(400, f"Role must be one of: {', '.join(sorted(valid_roles))}")
    if len(body.password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters.")

    email = body.email.strip().lower()
    if fetchone("SELECT id FROM profiles WHERE email = ?", (email,)):
        raise HTTPException(409, "An account with this email already exists.")

    uid = new_id()
    execute(
        "INSERT INTO profiles (id, email, name, role, approved, password_hash) VALUES (?,?,?,?,1,?)",
        (uid, email, body.name.strip(), body.role, _hash_password(body.password)),
    )
    return {"id": uid, "email": email, "name": body.name.strip(), "role": body.role}


# ── Admin: create single student account ─────────────────────────────────────

@router.post("/admin/create-student")
def create_student(body: CreateStudentRequest, user: dict = Depends(verify_token)):
    require_admin(user)
    if len(body.password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters.")

    email = body.email.strip().lower()
    usn = body.usn.strip().upper() if body.usn and body.usn.strip() else None
    if fetchone("SELECT id FROM profiles WHERE email = ?", (email,)):
        raise HTTPException(409, "An account with this email already exists.")

    uid = new_id()
    execute(
        "INSERT INTO profiles (id, email, name, usn, role, approved, password_hash) VALUES (?,?,?,?,?,1,?)",
        (uid, email, body.name.strip() or None, usn, "student", _hash_password(body.password)),
    )
    return {"id": uid, "email": email, "name": body.name, "usn": usn}


# ── Admin: preview xlsx (no DB writes) ───────────────────────────────────────

@router.post("/admin/preview-students")
async def preview_students(
    file: UploadFile = File(...),
    user: dict = Depends(verify_token),
):
    require_admin(user)
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are accepted.")
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not parse file: {exc}")

    ws = wb.active
    preview = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        cells = [str(v).strip() if v is not None else "" for v in row]
        while len(cells) < 5:
            cells.append("")
        sl_no, usn, name, email, password = cells[:5]
        preview.append({
            "sl_no": sl_no,
            "usn": usn,
            "name": name,
            "email": email,
            "password_set": bool(password),
        })
    return {"count": len(preview), "rows": preview}


# ── Admin: bulk student upload ────────────────────────────────────────────────

@router.post("/admin/upload-students")
async def upload_students(
    file: UploadFile = File(...),
    user: dict = Depends(verify_token),
):
    require_admin(user)
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are accepted.")
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not parse file: {exc}")

    ws = wb.active
    created, skipped, errors = [], [], []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        cells = [str(v).strip() if v is not None else "" for v in row]
        while len(cells) < 5:
            cells.append("")
        _sl, usn, name, email, password = cells[:5]

        if not email or not password:
            errors.append({"row": i, "error": "Email and password are required"})
            continue
        if len(password) < 6:
            errors.append({"row": i, "email": email, "error": "Password must be at least 6 characters"})
            continue

        email_lower = email.lower()
        if fetchone("SELECT id FROM profiles WHERE email = ?", (email_lower,)):
            skipped.append({"email": email, "reason": "already exists"})
            continue

        try:
            uid = new_id()
            execute(
                "INSERT INTO profiles (id, email, name, usn, role, approved, password_hash) VALUES (?,?,?,?,?,1,?)",
                (uid, email_lower, name or None, usn.upper() if usn else None, "student", _hash_password(password)),
            )
            created.append({"email": email, "name": name, "usn": usn})
        except Exception as e:
            errors.append({"row": i, "email": email, "error": str(e)})

    return {"created": len(created), "skipped": len(skipped), "errors": errors, "accounts": created}


# ── Admin: delete account ─────────────────────────────────────────────────────

@router.delete("/admin/users/{user_id}")
def delete_user(user_id: str, user: dict = Depends(verify_token)):
    require_admin(user)
    if user.get("sub") == user_id:
        raise HTTPException(400, "You cannot delete your own account.")
    if not fetchone("SELECT id FROM profiles WHERE id = ?", (user_id,)):
        raise HTTPException(404, "User not found.")
    execute("DELETE FROM profiles WHERE id = ?", (user_id,))
    return {"message": "Account deleted."}
