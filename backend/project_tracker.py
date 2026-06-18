"""
Project Tracker — domain API (SQLite/local version).
"""
from __future__ import annotations

import os
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel

from database import fetchall, fetchone, execute, executemany, new_id
from jwt_auth import (
    verify_token,
    require_admin,
    require_coordinator_for_course,
    require_coordinator_for_project,
    require_guide_for_project,
)

router = APIRouter(prefix="/api/tracker", tags=["project-tracker"])

UPLOAD_DIR = os.getenv("UPLOAD_DIR", "/data/uploads")

MAX_TEAM = 4
_TEAM_COLS: list[str] = []
for _i in range(1, MAX_TEAM + 1):
    _TEAM_COLS += [f"student_{_i}_usn", f"student_{_i}_name"]

TEMPLATE_COLUMNS: list[str] = ["project_title", "github_link", "guide_name", *_TEAM_COLS]
COURSE_TEMPLATE_COLUMNS: list[str] = ["team_no", "project_title", "github_url", "usn", "student_name"]

_ELEVATED = frozenset({"hod", "admin"})

REVIEW_META: dict = {
    1: {
        "phase": "Problem & Design Phase",
        "max_each": 10,
        "max_total": 50,
        "criteria": [
            "Problem Definition",
            "Literature Survey",
            "Proposed Solution",
            "Methodology Understanding",
            "Team Work & Presentation (Q&A)",
        ],
    },
    2: {
        "phase": "Implementation Phase",
        "max_each": 10,
        "max_total": 50,
        "criteria": [
            "Implementation Progress",
            "Methodology Application",
            "Verification & Validation",
            "Project Management",
            "Team Work & Presentation (Q&A)",
        ],
    },
    3: {
        "phase": "Demonstration Phase",
        "max_each": 25,
        "max_total": 100,
        "criteria": [
            "Effective Demonstration",
            "Test Case Coverage",
            "Completeness of Project as per Set Objectives",
            "Team Work & Presentation (Q&A)",
        ],
    },
}


# ── Pydantic schemas ──────────────────────────────────────────────────────────

class BatchIn(BaseModel):
    batch_name: str
    year: Optional[int] = None

class SemesterIn(BaseModel):
    sem_number: int

class CourseIn(BaseModel):
    course_name: str
    course_code: Optional[str] = None

class ProjectIn(BaseModel):
    title: Optional[str] = None
    github: Optional[str] = None
    guide: Optional[str] = None
    team_number: Optional[int] = None

class ProjectUpdate(BaseModel):
    title: Optional[str] = None
    github: Optional[str] = None
    guide: Optional[str] = None

class TeamIn(BaseModel):
    team_number: Optional[int] = None
    title: str
    github: Optional[str] = None
    students: Optional[list] = []

class StudentUpdate(BaseModel):
    usn: Optional[str] = None
    name: Optional[str] = None

class SemesterStatusIn(BaseModel):
    project_status: Optional[str] = None
    project_active: Optional[bool] = None

class ReviewIn(BaseModel):
    title: str
    date: str
    description: Optional[str] = None
    document_url: Optional[str] = None

class ReviewUpdate(BaseModel):
    title: Optional[str] = None
    date: Optional[str] = None
    description: Optional[str] = None
    document_url: Optional[str] = None

class CoordinatorIn(BaseModel):
    user_id: str

class AssignCoordinatorIn(BaseModel):
    user_id: str
    course_id: str

class GuideAssignIn(BaseModel):
    user_id: Optional[str] = None

class ExaminerIn(BaseModel):
    user_id: str

class EvalLockIn(BaseModel):
    is_locked: bool

class EvalMarkIn(BaseModel):
    student_id: str
    c1: Optional[int] = None
    c2: Optional[int] = None
    c3: Optional[int] = None
    c4: Optional[int] = None
    c5: Optional[int] = None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _create_eval_reviews(course_id: str) -> None:
    for n in (1, 2, 3):
        execute(
            "INSERT INTO evaluation_review (id, course_id, review_number, is_locked) VALUES (?,?,?,1) "
            "ON CONFLICT(course_id, review_number) DO NOTHING",
            (new_id(), course_id, n),
        )


def _enrich_guides(projects: list[dict]) -> None:
    guide_ids = list({p["guide_id"] for p in projects if p.get("guide_id")})
    if not guide_ids:
        for p in projects:
            p["guide_profile"] = None
        return
    ph = ",".join("?" * len(guide_ids))
    profiles = fetchall(f"SELECT id, name, email FROM profiles WHERE id IN ({ph})", tuple(guide_ids))
    pm = {p["id"]: p for p in profiles}
    for proj in projects:
        gid = proj.get("guide_id")
        proj["guide_profile"] = pm.get(gid) if gid else None


def _enrich_examiners(projects: list[dict]) -> None:
    project_ids = [p["project_id"] for p in projects]
    if not project_ids:
        for p in projects:
            p["examiners"] = []
        return
    ph = ",".join("?" * len(project_ids))
    rows = fetchall(
        f"SELECT project_id, user_id FROM project_examiner WHERE project_id IN ({ph})",
        tuple(project_ids),
    )
    user_ids = list({r["user_id"] for r in rows})
    prof_map: dict = {}
    if user_ids:
        uph = ",".join("?" * len(user_ids))
        profs = fetchall(f"SELECT id, name, email FROM profiles WHERE id IN ({uph})", tuple(user_ids))
        prof_map = {p["id"]: p for p in profs}
    by_project: dict = {}
    for r in rows:
        pid = r["project_id"]
        prof = prof_map.get(r["user_id"], {})
        by_project.setdefault(pid, []).append({
            "user_id": r["user_id"],
            "name": prof.get("name"),
            "email": prof.get("email"),
        })
    for p in projects:
        p["examiners"] = by_project.get(p["project_id"], [])


def _fetch_project_list(where_sql: str, params: tuple = ()) -> list[dict]:
    """Fetch projects + students + guide profiles + examiners matching where_sql."""
    rows = fetchall(f"""
        SELECT project_id, title, github, guide, guide_id, team_number, course_id
        FROM project
        {where_sql}
        ORDER BY CASE WHEN team_number IS NULL THEN 1 ELSE 0 END, team_number
    """, params)
    if not rows:
        return rows
    project_ids = [r["project_id"] for r in rows]
    ph = ",".join("?" * len(project_ids))
    students = fetchall(
        f"SELECT student_id, project_id, usn, name FROM student WHERE project_id IN ({ph})",
        tuple(project_ids),
    )
    by_proj: dict = {}
    for s in students:
        by_proj.setdefault(s["project_id"], []).append(
            {"student_id": s["student_id"], "usn": s["usn"], "name": s["name"]}
        )
    for r in rows:
        r["students"] = by_proj.get(r["project_id"], [])
    _enrich_guides(rows)
    _enrich_examiners(rows)
    return rows


def _enrich_eval_meta(rows: list[dict]) -> None:
    for r in rows:
        meta = REVIEW_META.get(r["review_number"], {})
        r["phase"]     = meta.get("phase", "")
        r["max_total"] = meta.get("max_total", 0)
        r["max_each"]  = meta.get("max_each", 10)
        r["criteria"]  = meta.get("criteria", [])


# ── Read endpoints ────────────────────────────────────────────────────────────

@router.get("/batches")
def list_batches(_user=Depends(verify_token)):
    batches = fetchall("SELECT batch_id, batch_name, year FROM batch ORDER BY year DESC")
    for b in batches:
        sems = fetchall(
            "SELECT project_status, project_active FROM semester WHERE batch_id = ?",
            (b["batch_id"],),
        )
        for s in sems:
            s["project_active"] = bool(s["project_active"])
        b["semester"] = sems
    return batches


@router.delete("/batches/{batch_id}")
def delete_batch(batch_id: str, user=Depends(verify_token)):
    require_admin(user)
    if not fetchone("SELECT batch_id FROM batch WHERE batch_id = ?", (batch_id,)):
        raise HTTPException(404, "Batch not found.")
    execute("DELETE FROM batch WHERE batch_id = ?", (batch_id,))
    return {"message": "Batch deleted."}


@router.get("/batches/{batch_id}/semesters")
def list_semesters(batch_id: str, _user=Depends(verify_token)):
    rows = fetchall(
        "SELECT * FROM semester WHERE batch_id = ? ORDER BY sem_number", (batch_id,)
    )
    for r in rows:
        r["project_active"] = bool(r["project_active"])
    return rows


@router.get("/me")
def get_me(user: dict = Depends(verify_token)):
    uid = user.get("sub")
    prof = fetchone(
        "SELECT id, email, name, usn, role, approved FROM profiles WHERE id = ?", (uid,)
    )
    if not prof:
        raise HTTPException(404, "Profile not found. Contact an admin to set up your account.")
    return prof


@router.get("/teachers")
def list_teachers(_user=Depends(verify_token)):
    return fetchall(
        "SELECT id, name, email, role FROM profiles "
        "WHERE role IN ('teacher','hod','admin') AND approved = 1 ORDER BY name"
    )


@router.patch("/semesters/{semester_id}/status")
def update_semester_status(semester_id: str, body: SemesterStatusIn, user=Depends(verify_token)):
    uid = user.get("sub")
    if not uid:
        raise HTTPException(401, "No subject claim in token.")
    role_row = fetchone("SELECT role FROM profiles WHERE id = ?", (uid,))
    user_role = role_row.get("role", "") if role_row else ""
    if user_role not in _ELEVATED:
        courses = fetchall("SELECT course_id FROM course WHERE semester_id = ?", (semester_id,))
        if not courses:
            raise HTTPException(403, "Not authorized.")
        course_ids = [c["course_id"] for c in courses]
        ph = ",".join("?" * len(course_ids))
        hit = fetchone(
            f"SELECT 1 FROM coordinator WHERE user_id = ? AND course_id IN ({ph})",
            (uid, *course_ids),
        )
        if not hit:
            raise HTTPException(403, "Only a course coordinator or admin may update semester status.")
    if body.project_status not in (None, "mini", "major"):
        raise HTTPException(400, "project_status must be 'mini', 'major', or null.")
    if body.project_status is None:
        execute(
            "UPDATE semester SET project_status = NULL, project_active = 0 WHERE semester_id = ?",
            (semester_id,),
        )
    elif body.project_active is not None:
        execute(
            "UPDATE semester SET project_status = ?, project_active = ? WHERE semester_id = ?",
            (body.project_status, int(body.project_active), semester_id),
        )
    else:
        execute(
            "UPDATE semester SET project_status = ? WHERE semester_id = ?",
            (body.project_status, semester_id),
        )
    row = fetchone("SELECT * FROM semester WHERE semester_id = ?", (semester_id,))
    if not row:
        raise HTTPException(404, "Semester not found.")
    row["project_active"] = bool(row["project_active"])
    return row


@router.delete("/semesters/{semester_id}/projects")
def clear_semester_projects(semester_id: str, user=Depends(verify_token)):
    require_admin(user)
    courses = fetchall("SELECT course_id FROM course WHERE semester_id = ?", (semester_id,))
    if not courses:
        return {"deleted": 0}
    course_ids = [c["course_id"] for c in courses]
    ph = ",".join("?" * len(course_ids))
    cur = execute(f"DELETE FROM project WHERE course_id IN ({ph})", tuple(course_ids))
    return {"deleted": cur.rowcount}


@router.delete("/semesters/{semester_id}")
def delete_semester(semester_id: str, user=Depends(verify_token)):
    require_admin(user)
    if not fetchone("SELECT semester_id FROM semester WHERE semester_id = ?", (semester_id,)):
        raise HTTPException(404, "Semester not found.")
    execute("DELETE FROM semester WHERE semester_id = ?", (semester_id,))
    return {"message": "Semester deleted."}


@router.get("/semesters/{semester_id}/courses")
def list_courses(semester_id: str, _user=Depends(verify_token)):
    return fetchall(
        "SELECT course_id, course_name, course_code, semester_id FROM course "
        "WHERE semester_id = ? ORDER BY course_name",
        (semester_id,),
    )


@router.delete("/courses/{course_id}")
def delete_course(course_id: str, user=Depends(verify_token)):
    require_admin(user)
    if not fetchone("SELECT course_id FROM course WHERE course_id = ?", (course_id,)):
        raise HTTPException(404, "Course not found.")
    execute("DELETE FROM course WHERE course_id = ?", (course_id,))
    return {"message": "Course deleted."}


@router.get("/courses/{course_id}/is-coordinator")
def is_coordinator_check(course_id: str, user=Depends(verify_token)):
    uid = user.get("sub")
    if not uid:
        return {"is_coordinator": False}
    prof = fetchone("SELECT role FROM profiles WHERE id = ?", (uid,))
    role = prof.get("role", "") if prof else ""
    if role in _ELEVATED:
        return {"is_coordinator": True}
    hit = fetchone(
        "SELECT 1 FROM coordinator WHERE user_id = ? AND course_id = ?", (uid, course_id)
    )
    return {"is_coordinator": bool(hit)}


@router.get("/courses/{course_id}/projects")
def list_projects(course_id: str, _user=Depends(verify_token)):
    return _fetch_project_list("WHERE course_id = ?", (course_id,))


@router.get("/courses/{course_id}/my-guide-teams")
def my_guide_teams(course_id: str, user=Depends(verify_token)):
    uid = user.get("sub")
    return _fetch_project_list("WHERE course_id = ? AND guide_id = ?", (course_id, uid))


@router.get("/courses/{course_id}/my-exam-teams")
def my_exam_teams(course_id: str, user=Depends(verify_token)):
    uid = user.get("sub")
    exam_rows = fetchall("SELECT project_id FROM project_examiner WHERE user_id = ?", (uid,))
    project_ids = [r["project_id"] for r in exam_rows]
    if not project_ids:
        return []
    ph = ",".join("?" * len(project_ids))
    return _fetch_project_list(
        f"WHERE course_id = ? AND project_id IN ({ph})", (course_id, *project_ids)
    )


@router.get("/projects/{project_id}")
def get_project(project_id: str, _user=Depends(verify_token)):
    row = fetchone("""
        SELECT p.project_id, p.title, p.github, p.guide, p.guide_id, p.team_number, p.course_id,
               c.course_id AS c_id, c.course_name, c.semester_id AS c_sem_id,
               s.semester_id AS s_id, s.sem_number, s.batch_id AS s_batch_id,
               b.batch_id AS b_id, b.batch_name, b.year AS b_year
        FROM project p
        LEFT JOIN course c ON c.course_id = p.course_id
        LEFT JOIN semester s ON s.semester_id = c.semester_id
        LEFT JOIN batch b ON b.batch_id = s.batch_id
        WHERE p.project_id = ?
    """, (project_id,))
    if not row:
        raise HTTPException(404, "Project not found.")
    students = fetchall(
        "SELECT student_id, usn, name FROM student WHERE project_id = ?", (project_id,)
    )
    result = {
        "project_id": row["project_id"],
        "title": row["title"],
        "github": row["github"],
        "guide": row["guide"],
        "guide_id": row["guide_id"],
        "team_number": row["team_number"],
        "course_id": row["course_id"],
        "students": students,
        "course": {
            "course_id": row["c_id"],
            "course_name": row["course_name"],
            "semester_id": row["c_sem_id"],
            "semester": {
                "semester_id": row["s_id"],
                "sem_number": row["sem_number"],
                "batch_id": row["s_batch_id"],
                "batch": {
                    "batch_id": row["b_id"],
                    "batch_name": row["batch_name"],
                    "year": row["b_year"],
                },
            },
        },
    }
    _enrich_guides([result])
    _enrich_examiners([result])
    return result


# ── Write endpoints ───────────────────────────────────────────────────────────

@router.post("/batches")
def create_batch(body: BatchIn, user=Depends(verify_token)):
    require_admin(user)
    bid = new_id()
    try:
        execute(
            "INSERT INTO batch (batch_id, batch_name, year) VALUES (?,?,?)",
            (bid, body.batch_name, body.year),
        )
    except Exception as exc:
        raise HTTPException(400, str(exc))
    return fetchone("SELECT * FROM batch WHERE batch_id = ?", (bid,))


@router.post("/batches/{batch_id}/semesters")
def create_semester(batch_id: str, body: SemesterIn, user=Depends(verify_token)):
    require_admin(user)
    existing = fetchone(
        "SELECT * FROM semester WHERE batch_id = ? AND sem_number = ?",
        (batch_id, body.sem_number),
    )
    if existing:
        existing["project_active"] = bool(existing["project_active"])
        return existing
    sid = new_id()
    try:
        execute(
            "INSERT INTO semester (semester_id, batch_id, sem_number) VALUES (?,?,?)",
            (sid, batch_id, body.sem_number),
        )
    except Exception as exc:
        raise HTTPException(400, str(exc))
    row = fetchone("SELECT * FROM semester WHERE semester_id = ?", (sid,))
    row["project_active"] = bool(row["project_active"])
    return row


@router.post("/semesters/{semester_id}/courses")
def create_course(semester_id: str, body: CourseIn, user=Depends(verify_token)):
    require_admin(user)
    cid = new_id()
    try:
        execute(
            "INSERT INTO course (course_id, semester_id, course_name, course_code) VALUES (?,?,?,?)",
            (cid, semester_id, body.course_name, body.course_code),
        )
        _create_eval_reviews(cid)
    except Exception as exc:
        raise HTTPException(400, str(exc))
    return fetchone("SELECT * FROM course WHERE course_id = ?", (cid,))


@router.post("/courses/{course_id}/coordinators")
def add_coordinator(course_id: str, body: CoordinatorIn, user=Depends(verify_token)):
    require_admin(user)
    execute(
        "INSERT OR IGNORE INTO coordinator (user_id, course_id) VALUES (?,?)",
        (body.user_id, course_id),
    )
    return fetchone(
        "SELECT user_id, course_id FROM coordinator WHERE user_id = ? AND course_id = ?",
        (body.user_id, course_id),
    )


@router.get("/courses/{course_id}/coordinators")
def list_coordinators(course_id: str, _user=Depends(verify_token)):
    return fetchall(
        "SELECT user_id, course_id FROM coordinator WHERE course_id = ?", (course_id,)
    )


@router.post("/courses/{course_id}/projects")
def create_project(course_id: str, body: ProjectIn, user=Depends(verify_token)):
    uid = user.get("sub")
    prof = fetchone("SELECT role FROM profiles WHERE id = ?", (uid,))
    if not prof or prof.get("role") == "student":
        raise HTTPException(403, "Students cannot create projects.")
    pid = new_id()
    fields = {k: v for k, v in body.model_dump().items() if v is not None}
    cols = ["project_id", "course_id"] + list(fields.keys())
    vals = [pid, course_id] + list(fields.values())
    ph = ",".join("?" * len(vals))
    execute(f"INSERT INTO project ({','.join(cols)}) VALUES ({ph})", tuple(vals))
    return fetchone("SELECT * FROM project WHERE project_id = ?", (pid,))


@router.post("/courses/{course_id}/teams")
def create_team(course_id: str, body: TeamIn, user=Depends(verify_token)):
    require_coordinator_for_course(course_id, user)
    pid = new_id()
    cols = ["project_id", "course_id", "title"]
    vals: list = [pid, course_id, body.title]
    if body.team_number is not None:
        cols.append("team_number"); vals.append(body.team_number)
    if body.github:
        cols.append("github"); vals.append(body.github)
    ph = ",".join("?" * len(vals))
    execute(f"INSERT INTO project ({','.join(cols)}) VALUES ({ph})", tuple(vals))
    if body.students:
        rows = [
            (new_id(), pid, s["usn"], s["name"])
            for s in body.students
            if isinstance(s, dict) and s.get("usn") and s.get("name")
        ]
        if rows:
            executemany(
                "INSERT INTO student (student_id, project_id, usn, name) VALUES (?,?,?,?)", rows
            )
    result = fetchone(
        "SELECT project_id, title, github, guide_id, team_number FROM project WHERE project_id = ?",
        (pid,),
    )
    result["students"] = fetchall(
        "SELECT student_id, usn, name FROM student WHERE project_id = ?", (pid,)
    )
    result["guide_profile"] = None
    result["examiners"] = []
    return result


@router.put("/projects/{project_id}")
def update_project(project_id: str, body: ProjectUpdate, user=Depends(verify_token)):
    require_guide_for_project(project_id, user)
    patch = {k: v for k, v in body.model_dump().items() if v is not None}
    if not patch:
        raise HTTPException(400, "No fields provided.")
    set_clause = ", ".join(f"{k} = ?" for k in patch)
    execute(
        f"UPDATE project SET {set_clause} WHERE project_id = ?",
        (*patch.values(), project_id),
    )
    row = fetchone("SELECT * FROM project WHERE project_id = ?", (project_id,))
    if not row:
        raise HTTPException(404, "Project not found.")
    return row


@router.delete("/projects/{project_id}")
def delete_project(project_id: str, user=Depends(verify_token)):
    require_guide_for_project(project_id, user)
    if not fetchone("SELECT project_id FROM project WHERE project_id = ?", (project_id,)):
        raise HTTPException(404, "Project not found.")
    execute("DELETE FROM project WHERE project_id = ?", (project_id,))
    return {"message": "Project deleted."}


@router.patch("/projects/{project_id}/guide")
def assign_guide(project_id: str, body: GuideAssignIn, user=Depends(verify_token)):
    require_coordinator_for_project(project_id, user)
    execute(
        "UPDATE project SET guide_id = ? WHERE project_id = ?", (body.user_id, project_id)
    )
    row = fetchone("SELECT * FROM project WHERE project_id = ?", (project_id,))
    if not row:
        raise HTTPException(404, "Project not found.")
    return row


@router.post("/projects/{project_id}/examiners")
def add_examiner(project_id: str, body: ExaminerIn, user=Depends(verify_token)):
    require_coordinator_for_project(project_id, user)
    existing = fetchall("SELECT id FROM project_examiner WHERE project_id = ?", (project_id,))
    if len(existing) >= 2:
        raise HTTPException(400, "Maximum 2 examiners allowed per project.")
    try:
        eid = new_id()
        execute(
            "INSERT INTO project_examiner (id, project_id, user_id) VALUES (?,?,?)",
            (eid, project_id, body.user_id),
        )
        return fetchone("SELECT * FROM project_examiner WHERE id = ?", (eid,))
    except Exception as exc:
        raise HTTPException(400, str(exc))


@router.delete("/projects/{project_id}/examiners/{user_id}")
def remove_examiner(project_id: str, user_id: str, user=Depends(verify_token)):
    require_coordinator_for_project(project_id, user)
    execute(
        "DELETE FROM project_examiner WHERE project_id = ? AND user_id = ?",
        (project_id, user_id),
    )
    return {"message": "Examiner removed."}


@router.put("/students/{student_id}")
def update_student(student_id: str, body: StudentUpdate, user=Depends(verify_token)):
    s = fetchone("SELECT project_id FROM student WHERE student_id = ?", (student_id,))
    if not s:
        raise HTTPException(404, "Student not found.")
    require_guide_for_project(s["project_id"], user)
    patch = {k: v for k, v in body.model_dump().items() if v is not None}
    if not patch:
        raise HTTPException(400, "No fields provided.")
    set_clause = ", ".join(f"{k} = ?" for k in patch)
    execute(
        f"UPDATE student SET {set_clause} WHERE student_id = ?",
        (*patch.values(), student_id),
    )
    return fetchone("SELECT * FROM student WHERE student_id = ?", (student_id,))


@router.delete("/students/{student_id}")
def delete_student(student_id: str, user=Depends(verify_token)):
    s = fetchone("SELECT project_id FROM student WHERE student_id = ?", (student_id,))
    if not s:
        raise HTTPException(404, "Student not found.")
    require_guide_for_project(s["project_id"], user)
    execute("DELETE FROM student WHERE student_id = ?", (student_id,))
    return {"message": "Student deleted."}


# ── Reviews ───────────────────────────────────────────────────────────────────

@router.get("/courses/{course_id}/reviews")
def list_reviews(course_id: str, _user=Depends(verify_token)):
    return fetchall("SELECT * FROM review WHERE course_id = ? ORDER BY date", (course_id,))


@router.post("/courses/{course_id}/reviews")
def create_review(course_id: str, body: ReviewIn, user=Depends(verify_token)):
    require_coordinator_for_course(course_id, user)
    uid = user.get("sub")
    rid = new_id()
    execute(
        "INSERT INTO review (review_id, course_id, scheduled_by, title, date, description, document_url) "
        "VALUES (?,?,?,?,?,?,?)",
        (rid, course_id, uid, body.title, body.date, body.description, body.document_url),
    )
    return fetchone("SELECT * FROM review WHERE review_id = ?", (rid,))


@router.put("/reviews/{review_id}")
def update_review(review_id: str, body: ReviewUpdate, user=Depends(verify_token)):
    rev = fetchone("SELECT course_id FROM review WHERE review_id = ?", (review_id,))
    if not rev:
        raise HTTPException(404, "Review not found.")
    require_coordinator_for_course(rev["course_id"], user)
    patch = {k: v for k, v in body.model_dump().items() if v is not None}
    if not patch:
        raise HTTPException(400, "No fields provided.")
    set_clause = ", ".join(f"{k} = ?" for k in patch)
    execute(
        f"UPDATE review SET {set_clause} WHERE review_id = ?",
        (*patch.values(), review_id),
    )
    return fetchone("SELECT * FROM review WHERE review_id = ?", (review_id,))


@router.delete("/reviews/{review_id}")
def delete_review(review_id: str, user=Depends(verify_token)):
    rev = fetchone("SELECT course_id FROM review WHERE review_id = ?", (review_id,))
    if not rev:
        raise HTTPException(404, "Review not found.")
    require_coordinator_for_course(rev["course_id"], user)
    execute("DELETE FROM review WHERE review_id = ?", (review_id,))
    return {"message": "Review deleted."}


# ── Course-level Excel upload ─────────────────────────────────────────────────

@router.post("/courses/{course_id}/upload-excel")
async def upload_course_excel(
    course_id: str,
    file: UploadFile = File(...),
    user=Depends(verify_token),
):
    require_coordinator_for_course(course_id, user)
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are accepted.")
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not parse the file ({exc}).")
    ws = wb.active
    header = [
        str(c.value).strip().lower().replace(" ", "_") if c.value is not None else ""
        for c in ws[1]
    ]
    expected = [col.lower() for col in COURSE_TEMPLATE_COLUMNS]
    for idx, (got, want) in enumerate(zip(header[: len(expected)], expected), start=1):
        if got != want:
            raise HTTPException(
                400,
                f"Column {idx} must be '{COURSE_TEMPLATE_COLUMNS[idx - 1]}' "
                f"but got '{ws[1][idx - 1].value}'. Re-download the template.",
            )
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    teams: dict = {}
    team_order: list = []
    current_no: Optional[str] = None
    for row in data_rows:
        if len(row) < 5:
            continue
        team_no_val = _str(row[0])
        title_val   = _str(row[1])
        github_val  = _str(row[2])
        usn_val     = _str(row[3])
        name_val    = _str(row[4])
        if team_no_val:
            current_no = team_no_val
            if current_no not in teams:
                try:
                    team_num: Optional[int] = int(team_no_val)
                except ValueError:
                    team_num = None
                teams[current_no] = {
                    "title": title_val or f"Team {team_no_val}",
                    "github": github_val,
                    "team_number": team_num,
                    "students": [],
                }
                team_order.append(current_no)
            else:
                if title_val:
                    teams[current_no]["title"] = title_val
                if github_val:
                    teams[current_no]["github"] = github_val
        if current_no and usn_val and name_val:
            teams[current_no]["students"].append({"usn": usn_val, "name": name_val})

    filled = [t for t in teams.values() if t["students"]]
    if not filled:
        raise HTTPException(400, "No teams with students found.")

    created = updated = total_students = 0
    for team_no_str in team_order:
        t = teams[team_no_str]
        if not t["students"]:
            continue
        team_num = t["team_number"]
        existing_id: Optional[str] = None
        if team_num is not None:
            ex = fetchone(
                "SELECT project_id FROM project WHERE course_id = ? AND team_number = ?",
                (course_id, team_num),
            )
            if ex:
                existing_id = ex["project_id"]
        if existing_id:
            execute(
                "UPDATE project SET title = ?, github = ? WHERE project_id = ?",
                (t["title"], t["github"], existing_id),
            )
            project_id = existing_id
            updated += 1
        else:
            project_id = new_id()
            cols = ["project_id", "course_id", "title"]
            vals: list = [project_id, course_id, t["title"]]
            if team_num is not None:
                cols.append("team_number"); vals.append(team_num)
            if t["github"]:
                cols.append("github"); vals.append(t["github"])
            ph = ",".join("?" * len(vals))
            execute(f"INSERT INTO project ({','.join(cols)}) VALUES ({ph})", tuple(vals))
            created += 1
        execute("DELETE FROM student WHERE project_id = ?", (project_id,))
        if t["students"]:
            student_rows = [
                (new_id(), project_id, s["usn"], s["name"]) for s in t["students"]
            ]
            executemany(
                "INSERT INTO student (student_id, project_id, usn, name) VALUES (?,?,?,?)",
                student_rows,
            )
            total_students += len(student_rows)
    return {"teams_created": created, "teams_updated": updated, "students_inserted": total_students}


@router.get("/courses/{course_id}/download-course-template")
def download_course_template(course_id: str, _user=Depends(verify_token)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"
    hdr_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    hdr_font = Font(color="F5A623", bold=True)
    ws.append(COURSE_TEMPLATE_COLUMNS)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    for team_num in range(1, 21):
        ws.append([team_num, "", "", "", ""])
        for _ in range(3):
            ws.append(["", "", "", "", ""])
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(width + 4, 18)
    iws = wb.create_sheet("Instructions")
    iws.append(["Column", "Required?", "Notes"])
    for cell in iws[1]:
        cell.font = Font(bold=True)
    for row in [
        ("team_no", "Yes (first row of team)", "Integer."),
        ("project_title", "Yes (first row of team)", "Defaults to 'Team N' if blank."),
        ("github_url", "Optional", "Can be added later inside the project."),
        ("usn", "Yes", "Student USN e.g. 4NM22AI001"),
        ("student_name", "Yes", "Student full name"),
    ]:
        iws.append(row)
    for col in iws.columns:
        width = max(len(str(c.value or "")) for c in col)
        iws.column_dimensions[col[0].column_letter].width = max(width + 4, 20)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=course_teams_template.xlsx"},
    )


@router.post("/projects/{project_id}/upload-excel")
async def upload_project_excel(
    project_id: str,
    file: UploadFile = File(...),
    user=Depends(verify_token),
):
    require_guide_for_project(project_id, user)
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are accepted.")
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not parse the file ({exc}).")
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if len(header) != len(TEMPLATE_COLUMNS):
        raise HTTPException(400, f"Header has {len(header)} columns; expected {len(TEMPLATE_COLUMNS)}.")
    for idx, (got, want) in enumerate(zip(header, TEMPLATE_COLUMNS), start=1):
        if got != want:
            raise HTTPException(400, f"Column {idx} must be '{want}' but got '{got}'.")
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    row = next((r for r in data_rows if any(v is not None and str(v).strip() for v in r)), None)
    if row is None:
        raise HTTPException(400, "Sheet has no data rows.")
    cells = dict(zip(TEMPLATE_COLUMNS, row))
    proj_patch = {
        k: _str(cells.get(src))
        for k, src in [("title", "project_title"), ("github", "github_link"), ("guide", "guide_name")]
        if _str(cells.get(src)) is not None
    }
    if proj_patch:
        set_clause = ", ".join(f"{k} = ?" for k in proj_patch)
        execute(
            f"UPDATE project SET {set_clause} WHERE project_id = ?",
            (*proj_patch.values(), project_id),
        )
    execute("DELETE FROM student WHERE project_id = ?", (project_id,))
    new_students: list = []
    skipped_slots: list = []
    for i in range(1, MAX_TEAM + 1):
        usn = _str(cells.get(f"student_{i}_usn"))
        name = _str(cells.get(f"student_{i}_name"))
        if usn and name:
            new_students.append((new_id(), project_id, usn, name))
        elif usn or name:
            skipped_slots.append(i)
    if new_students:
        executemany(
            "INSERT INTO student (student_id, project_id, usn, name) VALUES (?,?,?,?)",
            new_students,
        )
    return {
        "project": fetchone("SELECT * FROM project WHERE project_id = ?", (project_id,)),
        "students_inserted": len(new_students),
        "partial_slots_skipped": len(skipped_slots),
        "applied_project_fields": list(proj_patch.keys()),
    }


@router.get("/download-template")
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    hdr_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    hdr_font = Font(color="F5A623", bold=True)
    ws.append(TEMPLATE_COLUMNS)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    ws.append(["Image Classifier", "https://github.com/team/ml-proj", "Dr. Smith",
                "1CS21AI001", "Alice Kumar", "1CS21AI002", "Bob Nair", "1CS21AI003", "Carol Menon", "", ""])
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(width + 4, 16)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=project_template.xlsx"},
    )


@router.get("/projects/{project_id}/readme")
def get_project_readme(project_id: str, _user=Depends(verify_token)):
    from github_utils import parse_repo_url, fetch_readme
    proj = fetchone("SELECT github FROM project WHERE project_id = ?", (project_id,))
    if not proj:
        raise HTTPException(404, "Project not found.")
    parsed = parse_repo_url(proj.get("github"))
    if not parsed:
        return {"found": False, "content": "", "reason": "no-github-link"}
    content = fetch_readme(*parsed)
    if not content:
        return {"found": False, "content": "", "reason": "private-or-missing"}
    return {"found": True, "content": content}


# ── Admin endpoints ───────────────────────────────────────────────────────────

@router.get("/admin/users")
def admin_list_users(user: dict = Depends(verify_token)):
    require_admin(user)
    return fetchall(
        "SELECT id, email, name, usn, role, approved FROM profiles "
        "WHERE role IN ('teacher','hod','student') ORDER BY email"
    )


@router.get("/admin/courses")
def admin_list_all_courses(user: dict = Depends(verify_token)):
    require_admin(user)
    rows = fetchall("""
        SELECT c.course_id, c.course_name, c.course_code,
               s.sem_number, b.batch_name
        FROM course c
        LEFT JOIN semester s ON s.semester_id = c.semester_id
        LEFT JOIN batch b ON b.batch_id = s.batch_id
    """)
    return [
        {
            "course_id": r["course_id"],
            "course_name": r["course_name"],
            "course_code": r["course_code"],
            "semester": {
                "sem_number": r["sem_number"],
                "batch": {"batch_name": r["batch_name"]},
            },
        }
        for r in rows
    ]


@router.get("/admin/coordinators")
def admin_list_all_coordinators(user: dict = Depends(verify_token)):
    require_admin(user)
    coords = fetchall("SELECT user_id, course_id FROM coordinator")
    all_profiles = fetchall("SELECT id, email, role FROM profiles")
    pm = {p["id"]: p for p in all_profiles}
    all_courses = fetchall("SELECT course_id, course_name, course_code FROM course")
    cm = {c["course_id"]: c for c in all_courses}
    for c in coords:
        c["email"] = pm.get(c["user_id"], {}).get("email", "")
        course = cm.get(c["course_id"], {})
        c["course_name"] = course.get("course_name", "")
        c["course_code"] = course.get("course_code", "")
    return coords


@router.post("/admin/assign-coordinator")
def admin_assign_coordinator(body: AssignCoordinatorIn, user: dict = Depends(verify_token)):
    require_admin(user)
    execute(
        "INSERT OR IGNORE INTO coordinator (user_id, course_id) VALUES (?,?)",
        (body.user_id, body.course_id),
    )
    return fetchone(
        "SELECT user_id, course_id FROM coordinator WHERE user_id = ? AND course_id = ?",
        (body.user_id, body.course_id),
    )


@router.delete("/admin/coordinators/{course_id}/{user_id}")
def admin_remove_coordinator(course_id: str, user_id: str, user: dict = Depends(verify_token)):
    require_admin(user)
    execute(
        "DELETE FROM coordinator WHERE course_id = ? AND user_id = ?", (course_id, user_id)
    )
    return {"message": "Coordinator removed."}


# ── Evaluation reviews + marks ────────────────────────────────────────────────

@router.get("/courses/{course_id}/evaluation-reviews")
def list_eval_reviews(course_id: str, _user=Depends(verify_token)):
    rows = fetchall(
        "SELECT * FROM evaluation_review WHERE course_id = ? ORDER BY review_number",
        (course_id,),
    )
    for r in rows:
        r["is_locked"] = bool(r["is_locked"])
    _enrich_eval_meta(rows)
    return rows


@router.patch("/evaluation-reviews/{eval_review_id}/lock")
def toggle_eval_review_lock(eval_review_id: str, body: EvalLockIn, user=Depends(verify_token)):
    rev = fetchone("SELECT course_id FROM evaluation_review WHERE id = ?", (eval_review_id,))
    if not rev:
        raise HTTPException(404, "Evaluation review not found.")
    require_coordinator_for_course(rev["course_id"], user)
    execute(
        "UPDATE evaluation_review SET is_locked = ? WHERE id = ?",
        (int(body.is_locked), eval_review_id),
    )
    row = fetchone("SELECT * FROM evaluation_review WHERE id = ?", (eval_review_id,))
    row["is_locked"] = bool(row["is_locked"])
    return row


@router.get("/evaluation-reviews/{eval_review_id}/my-marks/{project_id}")
def get_my_eval_marks(eval_review_id: str, project_id: str, user=Depends(verify_token)):
    uid = user.get("sub")
    if not uid:
        raise HTTPException(401, "No subject claim.")
    return fetchall(
        "SELECT student_id, scorer_type, c1, c2, c3, c4, c5 FROM evaluation_mark "
        "WHERE eval_review_id = ? AND project_id = ? AND scorer_id = ?",
        (eval_review_id, project_id, uid),
    )


@router.post("/evaluation-reviews/{eval_review_id}/marks/{project_id}")
def submit_eval_marks(
    eval_review_id: str,
    project_id: str,
    body: list[EvalMarkIn],
    user=Depends(verify_token),
):
    uid = user.get("sub")
    if not uid:
        raise HTTPException(401, "No subject claim.")
    rev = fetchone(
        "SELECT is_locked, review_number FROM evaluation_review WHERE id = ?", (eval_review_id,)
    )
    if not rev:
        raise HTTPException(404, "Evaluation review not found.")
    if rev["is_locked"]:
        raise HTTPException(403, "This review is locked. Ask the coordinator to unlock it.")
    proj = fetchone("SELECT guide_id FROM project WHERE project_id = ?", (project_id,))
    if not proj:
        raise HTTPException(404, "Project not found.")
    is_guide = proj.get("guide_id") == uid
    is_examiner = bool(fetchone(
        "SELECT 1 FROM project_examiner WHERE project_id = ? AND user_id = ?",
        (project_id, uid),
    ))
    if not is_guide and not is_examiner:
        raise HTTPException(403, "You are not assigned as guide or examiner for this project.")
    scorer_type = "guide" if is_guide else "evaluator"
    meta = REVIEW_META.get(rev["review_number"], {})
    max_each = meta.get("max_each", 10)
    num_criteria = len(meta.get("criteria", []))
    now_ts = datetime.now(timezone.utc).isoformat()
    for entry in body:
        marks = [entry.c1, entry.c2, entry.c3, entry.c4, entry.c5]
        for i, m in enumerate(marks[:num_criteria]):
            if m is not None and not (0 <= m <= max_each):
                raise HTTPException(400, f"Criterion {i + 1} must be 0–{max_each}.")
        execute(
            """
            INSERT INTO evaluation_mark
                (id, eval_review_id, project_id, student_id, scorer_id, scorer_type,
                 c1, c2, c3, c4, c5, submitted_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(eval_review_id, student_id, scorer_id, scorer_type)
            DO UPDATE SET
                project_id   = excluded.project_id,
                c1 = excluded.c1, c2 = excluded.c2, c3 = excluded.c3,
                c4 = excluded.c4, c5 = excluded.c5,
                submitted_at = excluded.submitted_at
            """,
            (new_id(), eval_review_id, project_id, entry.student_id, uid, scorer_type,
             entry.c1, entry.c2, entry.c3, entry.c4, entry.c5, now_ts),
        )
    return {"submitted": len(body), "scorer_type": scorer_type}


@router.get("/evaluation-reviews/{eval_review_id}/summary")
def get_eval_review_summary(eval_review_id: str, user=Depends(verify_token)):
    uid = user.get("sub")
    if not uid:
        raise HTTPException(401, "No subject claim.")
    rev = fetchone(
        "SELECT course_id, review_number FROM evaluation_review WHERE id = ?", (eval_review_id,)
    )
    if not rev:
        raise HTTPException(404, "Evaluation review not found.")
    course_id     = rev["course_id"]
    review_number = rev["review_number"]
    prof = fetchone("SELECT role FROM profiles WHERE id = ?", (uid,))
    role = prof.get("role", "") if prof else ""
    is_coord = role in _ELEVATED or bool(fetchone(
        "SELECT 1 FROM coordinator WHERE user_id = ? AND course_id = ?", (uid, course_id)
    ))
    if not is_coord:
        guide_check = fetchall(
            "SELECT project_id FROM project WHERE course_id = ? AND guide_id = ?", (course_id, uid)
        )
        if not guide_check:
            raise HTTPException(403, "Access restricted to coordinators and guides.")
    if is_coord:
        proj_rows = fetchall(
            "SELECT project_id, title, team_number FROM project WHERE course_id = ? "
            "ORDER BY CASE WHEN team_number IS NULL THEN 1 ELSE 0 END, team_number",
            (course_id,),
        )
    else:
        proj_rows = fetchall(
            "SELECT project_id, title, team_number FROM project WHERE course_id = ? AND guide_id = ? "
            "ORDER BY CASE WHEN team_number IS NULL THEN 1 ELSE 0 END, team_number",
            (course_id, uid),
        )
    if not proj_rows:
        return []
    project_ids = [p["project_id"] for p in proj_rows]
    ph = ",".join("?" * len(project_ids))
    all_students = fetchall(
        f"SELECT student_id, project_id, usn, name FROM student WHERE project_id IN ({ph})",
        tuple(project_ids),
    )
    stu_by_proj: dict = {}
    for s in all_students:
        stu_by_proj.setdefault(s["project_id"], []).append(s)
    marks = fetchall(
        f"SELECT project_id, student_id, scorer_id, scorer_type, c1, c2, c3, c4, c5 "
        f"FROM evaluation_mark WHERE eval_review_id = ? AND project_id IN ({ph})",
        (eval_review_id, *project_ids),
    )
    scorer_ids = list({m["scorer_id"] for m in marks})
    scorer_map: dict = {}
    if scorer_ids:
        sph = ",".join("?" * len(scorer_ids))
        profs = fetchall(f"SELECT id, name, email FROM profiles WHERE id IN ({sph})", tuple(scorer_ids))
        scorer_map = {p["id"]: p for p in profs}
    meta  = REVIEW_META.get(review_number, {})
    num_c = len(meta.get("criteria", []))
    result = []
    for proj in proj_rows:
        pid = proj["project_id"]
        proj_marks = [m for m in marks if m["project_id"] == pid]
        students_out = []
        for s in stu_by_proj.get(pid, []):
            sid = s["student_id"]
            s_marks = [m for m in proj_marks if m["student_id"] == sid]
            marks_out = []
            for m in s_marks:
                sc    = scorer_map.get(m["scorer_id"], {})
                total = sum((m.get(f"c{i}") or 0) for i in range(1, num_c + 1))
                marks_out.append({
                    "scorer_name": sc.get("name") or sc.get("email", "Unknown"),
                    "scorer_type": m["scorer_type"],
                    **{f"c{i}": m.get(f"c{i}") for i in range(1, 6)},
                    "total": total,
                })
            students_out.append({**s, "marks": marks_out})
        result.append({
            "project_id":  pid,
            "title":       proj.get("title", ""),
            "team_number": proj.get("team_number"),
            "students":    students_out,
        })
    return result


@router.get("/courses/{course_id}/download-marks-excel")
def download_marks_excel(course_id: str, user=Depends(verify_token)):
    require_coordinator_for_course(course_id, user)

    course_row = fetchone(
        "SELECT course_name, course_code, semester_id FROM course WHERE course_id = ?", (course_id,)
    )
    if not course_row:
        raise HTTPException(404, "Course not found.")
    course_name  = course_row.get("course_name", "")
    course_code  = course_row.get("course_code", "") or ""
    course_label = f"{course_code} / {course_name}" if course_code else course_name

    sem_row    = fetchone(
        "SELECT sem_number, batch_id FROM semester WHERE semester_id = ?", (course_row["semester_id"],)
    )
    sem        = sem_row or {}
    sem_number = sem.get("sem_number", "")
    batch_row  = fetchone("SELECT year FROM batch WHERE batch_id = ?", (sem.get("batch_id", ""),))
    batch_year = batch_row.get("year") if batch_row else None
    acad_year  = f"{batch_year}-{str(batch_year + 1)[2:]}" if batch_year else ""

    projects = fetchall(
        "SELECT project_id, title, guide_id, team_number FROM project WHERE course_id = ? "
        "ORDER BY CASE WHEN team_number IS NULL THEN 1 ELSE 0 END, team_number",
        (course_id,),
    )
    if not projects:
        raise HTTPException(400, "No projects in this course.")

    project_ids = [p["project_id"] for p in projects]
    ph = ",".join("?" * len(project_ids))

    all_students = fetchall(
        f"SELECT student_id, project_id, usn, name FROM student WHERE project_id IN ({ph})",
        tuple(project_ids),
    )
    stu_by_proj: dict = {}
    for s in all_students:
        stu_by_proj.setdefault(s["project_id"], []).append(s)
    for p in projects:
        p["students"] = sorted(stu_by_proj.get(p["project_id"], []), key=lambda s: s.get("usn") or "")

    exam_rows = fetchall(
        f"SELECT project_id, user_id, created_at FROM project_examiner WHERE project_id IN ({ph})",
        tuple(project_ids),
    )
    exam_by_proj: dict = {}
    for e in sorted(exam_rows, key=lambda x: x.get("created_at") or ""):
        exam_by_proj.setdefault(e["project_id"], []).append(e["user_id"])

    scorer_ids = list(
        {p.get("guide_id") for p in projects if p.get("guide_id")}
        | {uid for uids in exam_by_proj.values() for uid in uids}
    )
    scorer_name_map: dict = {}
    if scorer_ids:
        sph = ",".join("?" * len(scorer_ids))
        profs = fetchall(f"SELECT id, name FROM profiles WHERE id IN ({sph})", tuple(scorer_ids))
        scorer_name_map = {p["id"]: p.get("name", "") for p in profs}

    eval_reviews = fetchall(
        "SELECT id, review_number FROM evaluation_review WHERE course_id = ? ORDER BY review_number",
        (course_id,),
    )
    rev_id_by_num = {r["review_number"]: r["id"] for r in eval_reviews}
    all_rev_ids = list(rev_id_by_num.values())
    all_marks: list = []
    if all_rev_ids:
        rph = ",".join("?" * len(all_rev_ids))
        all_marks = fetchall(
            f"SELECT eval_review_id, project_id, student_id, scorer_id, c1, c2, c3, c4, c5 "
            f"FROM evaluation_mark WHERE eval_review_id IN ({rph})",
            tuple(all_rev_ids),
        )
    marks_idx: dict = {}
    for m in all_marks:
        marks_idx[(m["eval_review_id"], m["project_id"], m["scorer_id"], m["student_id"])] = m

    THIN_SIDE   = Side(style="thin")
    ALL_BORDER  = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
    title_font  = Font(name="Calibri", size=14, bold=True)
    hdr_font    = Font(name="Calibri", size=9, bold=True)
    info_font   = Font(name="Calibri", size=9)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_mid  = Alignment(horizontal="center", vertical="center")
    right_mid   = Alignment(horizontal="right",  vertical="center")
    left_mid    = Alignment(horizontal="left",   vertical="center")

    def _col_letter(n: int) -> str:
        r = ""
        while n > 0:
            n, rem = divmod(n - 1, 26)
            r = chr(65 + rem) + r
        return r

    def _border_row(ws, row: int, start_col: int, end_col: int) -> None:
        for col in range(start_col, end_col + 1):
            left  = THIN_SIDE if col == start_col else Side(style=None)
            right = THIN_SIDE if col == end_col   else Side(style=None)
            ws.cell(row=row, column=col).border = Border(
                left=left, right=right, top=THIN_SIDE, bottom=THIN_SIDE
            )

    wb = Workbook()
    wb.remove(wb.active)

    def _make_sheet(rev_num: int, sheet_label: str, scorer_key):
        meta       = REVIEW_META[rev_num]
        criteria   = meta["criteria"]
        num_c      = len(criteria)
        max_each   = meta["max_each"]
        max_total  = meta["max_total"]
        rev_id     = rev_id_by_num.get(rev_num)
        total_cols = 3 + num_c + 1
        split_col  = total_cols - 2

        ws = wb.create_sheet(title=f"R{rev_num} {sheet_label}"[:31])
        ws.column_dimensions["A"].width = 7
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 24
        for i in range(num_c):
            ws.column_dimensions[_col_letter(4 + i)].width = 14
        ws.column_dimensions[_col_letter(4 + num_c)].width = 10

        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        tc = ws.cell(row=row, column=1, value=f"Review {rev_num} : {meta['phase']}")
        tc.font = title_font
        tc.alignment = center_wrap
        ws.row_dimensions[row].height = 28
        row += 1

        for proj in projects:
            pid      = proj["project_id"]
            students = proj.get("students") or []
            if scorer_key == "guide":
                scorer_id = proj.get("guide_id")
            else:
                exams     = exam_by_proj.get(pid, [])
                scorer_id = exams[scorer_key] if len(exams) > scorer_key else None
            scorer_name = scorer_name_map.get(scorer_id, "") if scorer_id else ""

            if row > 2:
                row += 1

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=split_col - 1)
            c = ws.cell(row=row, column=1, value=f"Course Code/Course Title:{course_label}")
            c.font = info_font; c.alignment = left_mid
            _border_row(ws, row, 1, split_col - 1)
            ws.merge_cells(start_row=row, start_column=split_col, end_row=row, end_column=total_cols)
            c = ws.cell(row=row, column=split_col, value=f"Academic Year:{acad_year}")
            c.font = info_font; c.alignment = right_mid
            _border_row(ws, row, split_col, total_cols)
            row += 1

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=split_col - 1)
            c = ws.cell(row=row, column=1, value=f"Semester:{sem_number}")
            c.font = info_font; c.alignment = left_mid
            _border_row(ws, row, 1, split_col - 1)
            ws.merge_cells(start_row=row, start_column=split_col, end_row=row, end_column=total_cols)
            c = ws.cell(row=row, column=split_col, value=f"Evaluator:{scorer_name}")
            c.font = info_font; c.alignment = right_mid
            _border_row(ws, row, split_col, total_cols)
            row += 1

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=split_col - 1)
            c = ws.cell(row=row, column=1, value=f"Project Title:{proj.get('title', '')}")
            c.font = info_font; c.alignment = left_mid
            _border_row(ws, row, 1, split_col - 1)
            ws.merge_cells(start_row=row, start_column=split_col, end_row=row, end_column=total_cols)
            c = ws.cell(row=row, column=split_col, value="Date:")
            c.font = info_font; c.alignment = right_mid
            _border_row(ws, row, split_col, total_cols)
            row += 1

            for col_idx, val in enumerate(["Sl. No", "USN", "Names"], 1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = hdr_font; c.border = ALL_BORDER; c.alignment = center_wrap
            for i, crit in enumerate(criteria):
                c = ws.cell(row=row, column=4 + i, value=f"{crit}\n({max_each} Marks)")
                c.font = hdr_font; c.border = ALL_BORDER; c.alignment = center_wrap
            c = ws.cell(row=row, column=4 + num_c, value=f"Total\n({max_total} Marks)")
            c.font = hdr_font; c.border = ALL_BORDER; c.alignment = center_wrap
            ws.row_dimensions[row].height = 44
            row += 1

            for sl, stu in enumerate(students, 1):
                stu_id   = stu["student_id"]
                mark_row = marks_idx.get((rev_id, pid, scorer_id, stu_id)) if (rev_id and scorer_id) else None
                ws.cell(row=row, column=1, value=sl).border        = ALL_BORDER
                ws.cell(row=row, column=1).alignment               = center_mid
                ws.cell(row=row, column=2, value=stu.get("usn", "")).border = ALL_BORDER
                ws.cell(row=row, column=2).alignment               = center_mid
                ws.cell(row=row, column=3, value=stu.get("name", "")).border = ALL_BORDER
                ws.cell(row=row, column=3).alignment               = left_mid
                total     = 0
                has_marks = False
                for i in range(num_c):
                    val = mark_row.get(f"c{i + 1}") if mark_row else None
                    c   = ws.cell(row=row, column=4 + i, value=val)
                    c.border = ALL_BORDER; c.alignment = center_mid
                    if val is not None:
                        total += val; has_marks = True
                c = ws.cell(row=row, column=4 + num_c, value=total if has_marks else None)
                c.border = ALL_BORDER; c.alignment = center_mid
                row += 1

    for rev_num in (1, 2, 3):
        _make_sheet(rev_num, "Guide", "guide")
        for eval_idx in (0, 1):
            if any(len(exam_by_proj.get(p["project_id"], [])) > eval_idx for p in projects):
                _make_sheet(rev_num, f"Eval {eval_idx + 1}", eval_idx)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = (course_code or course_name).replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="evaluation_{safe_name}.xlsx"'},
    )


@router.get("/courses/{course_id}/download-ce-sheet")
def download_ce_sheet(course_id: str, user=Depends(verify_token)):
    require_coordinator_for_course(course_id, user)

    course_row = fetchone(
        "SELECT course_name, course_code, semester_id FROM course WHERE course_id = ?", (course_id,)
    )
    if not course_row:
        raise HTTPException(404, "Course not found.")
    course_name  = course_row.get("course_name", "")
    course_code  = course_row.get("course_code", "") or ""
    course_label = f"{course_code} / {course_name}" if course_code else course_name

    sem_row    = fetchone(
        "SELECT sem_number, batch_id FROM semester WHERE semester_id = ?", (course_row["semester_id"],)
    )
    sem        = sem_row or {}
    sem_number = sem.get("sem_number", "")
    batch_row  = fetchone("SELECT year FROM batch WHERE batch_id = ?", (sem.get("batch_id", ""),))
    batch_year = batch_row.get("year") if batch_row else None
    acad_year  = f"{batch_year}-{str(batch_year + 1)[2:]}" if batch_year else ""

    projects = fetchall(
        "SELECT project_id, guide_id FROM project WHERE course_id = ? "
        "ORDER BY CASE WHEN team_number IS NULL THEN 1 ELSE 0 END, team_number",
        (course_id,),
    )
    project_ids = [p["project_id"] for p in projects]
    if not project_ids:
        raise HTTPException(400, "No students found in this course.")

    ph = ",".join("?" * len(project_ids))
    raw_students = fetchall(
        f"SELECT student_id, project_id, usn, name FROM student WHERE project_id IN ({ph})",
        tuple(project_ids),
    )
    guide_map = {p["project_id"]: p.get("guide_id") for p in projects}
    all_students = []
    for p in projects:
        pid = p["project_id"]
        for s in sorted(
            [x for x in raw_students if x["project_id"] == pid],
            key=lambda x: x.get("usn") or "",
        ):
            all_students.append({**s, "guide_id": guide_map.get(pid)})

    if not all_students:
        raise HTTPException(400, "No students found in this course.")

    exam_rows = fetchall(
        f"SELECT project_id, user_id FROM project_examiner WHERE project_id IN ({ph})",
        tuple(project_ids),
    )
    examiners_by_proj: dict = {}
    for e in exam_rows:
        examiners_by_proj.setdefault(e["project_id"], []).append(e["user_id"])

    eval_reviews = fetchall(
        "SELECT id, review_number FROM evaluation_review WHERE course_id = ? ORDER BY review_number",
        (course_id,),
    )
    rev_id_by_num = {r["review_number"]: r["id"] for r in eval_reviews}
    all_rev_ids = list(rev_id_by_num.values())
    raw_marks: list = []
    if all_rev_ids:
        rph = ",".join("?" * len(all_rev_ids))
        raw_marks = fetchall(
            f"SELECT eval_review_id, project_id, student_id, scorer_id, scorer_type, c1, c2, c3, c4, c5 "
            f"FROM evaluation_mark WHERE eval_review_id IN ({rph})",
            tuple(all_rev_ids),
        )
    marks_idx: dict = {}
    for m in raw_marks:
        marks_idx[(m["eval_review_id"], m["student_id"], m["scorer_id"])] = m

    def _total(mark_row, num_criteria: int):
        if not mark_row:
            return None
        vals = [mark_row.get(f"c{i}") for i in range(1, num_criteria + 1)]
        if all(v is None for v in vals):
            return None
        return sum(v or 0 for v in vals)

    def _student_typed_avg(student_id: str, project_id: str, rev_num: int, stype: str):
        rev_id = rev_id_by_num.get(rev_num)
        if not rev_id:
            return None
        num_c = len(REVIEW_META.get(rev_num, {}).get("criteria", []))
        if stype == "guide":
            gid = guide_map.get(project_id)
            if not gid:
                return None
            return _total(marks_idx.get((rev_id, student_id, gid)), num_c)
        else:
            totals = [
                t for eid in examiners_by_proj.get(project_id, [])
                if (t := _total(marks_idx.get((rev_id, student_id, eid)), num_c)) is not None
            ]
            return sum(totals) / len(totals) if totals else None

    def _avg2(a, b):
        vals = [x for x in [a, b] if x is not None]
        return sum(vals) / len(vals) if vals else None

    THIN     = Side(style="thin")
    ALL_BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    title_f  = Font(name="Calibri", size=13, bold=True)
    hdr_f    = Font(name="Calibri", size=9, bold=True)
    body_f   = Font(name="Calibri", size=9)
    c_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_align  = Alignment(horizontal="left",   vertical="center")

    wb  = Workbook()
    ws  = wb.active
    ws.title = "CE Marks"

    col_widths = [5, 14, 26, 14, 14, 14, 14, 14, 14, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    t = ws.cell(2, 1, "Mini Project Evaluation marks")
    t.font = title_f; t.alignment = c_align
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A3:J3")
    c = ws.cell(3, 1, f"Course Code/Course Title: {course_label}")
    c.font = body_f; c.alignment = l_align
    ws.merge_cells("A4:J4")
    c = ws.cell(4, 1, f"Semester: {sem_number}    Academic Year: {acad_year}")
    c.font = body_f; c.alignment = l_align

    headers_row5 = [
        "Sl. No", "USN", "Student Name",
        "Review 1-Guide\n(Problem & Design Phase)",
        "Review 1-Evaluator\n(Problem & Design Phase)",
        "Review 2-Guide\n(Implementation Phase)",
        "Review 2-Evaluator\n(Implementation Phase)",
        "Review 3-Guide\n(Demonstration Phase)",
        "Review 3-Evaluator\n(Demonstration Phase)",
        "Final CE",
    ]
    max_row5 = ["", "", "",
                "Max Marks-50", "Max Marks-50", "Max Marks-50",
                "Max Marks-50", "Max Marks-100", "Max Marks-100", "Max Marks-100"]
    hdr_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    sub_fill = PatternFill(start_color="1C2E40", end_color="1C2E40", fill_type="solid")
    hdr_fc   = Font(name="Calibri", size=9, bold=True, color="F5A623")
    sub_fc   = Font(name="Calibri", size=8, bold=False, color="AAAAAA")

    for col, (h, s) in enumerate(zip(headers_row5, max_row5), 1):
        c1 = ws.cell(5, col, h)
        c1.font = hdr_fc; c1.fill = hdr_fill; c1.border = ALL_BDR; c1.alignment = c_align
        c2 = ws.cell(6, col, s)
        c2.font = sub_fc; c2.fill = sub_fill; c2.border = ALL_BDR; c2.alignment = c_align
    ws.row_dimensions[5].height = 42
    ws.row_dimensions[6].height = 16

    ce_col = 10
    for sl, stu in enumerate(all_students, 1):
        row = 6 + sl
        sid  = stu["student_id"]
        pid  = stu["project_id"]
        r1_g = _student_typed_avg(sid, pid, 1, "guide")
        r1_e = _student_typed_avg(sid, pid, 1, "evaluator")
        r2_g = _student_typed_avg(sid, pid, 2, "guide")
        r2_e = _student_typed_avg(sid, pid, 2, "evaluator")
        r3_g = _student_typed_avg(sid, pid, 3, "guide")
        r3_e = _student_typed_avg(sid, pid, 3, "evaluator")
        r1_avg = _avg2(r1_g, r1_e)
        r2_avg = _avg2(r2_g, r2_e)
        r3_avg = _avg2(r3_g, r3_e)
        if r1_avg is not None or r2_avg is not None or r3_avg is not None:
            r1r2 = (r1_avg or 0) + (r2_avg or 0)
            r3   = r3_avg or 0
            ce   = round((r1r2 + r3) / 2)
        else:
            ce = None
        values = [sl, stu["usn"], stu["name"], r1_g, r1_e, r2_g, r2_e, r3_g, r3_e, ce]
        alt_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid") if sl % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row, col)
            if val is not None:
                cell.value = round(val) if isinstance(val, float) else val
            cell.font = body_f; cell.border = ALL_BDR
            cell.alignment = c_align if col != 3 else l_align
            if alt_fill:
                cell.fill = alt_fill
        ce_cell = ws.cell(row, ce_col)
        ce_cell.font = Font(name="Calibri", size=9, bold=True)
        if ce is not None:
            ce_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    footer_row = 6 + len(all_students) + 1
    ws.merge_cells(f"A{footer_row}:C{footer_row}")
    ws.cell(footer_row, 1, "Formula: CE = round((R1_avg + R2_avg + R3_avg) / 2)").font = Font(
        size=7, italic=True, color="888888"
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = (course_code or course_name).replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="CE_marks_{safe_name}.xlsx"'},
    )


# ── Course documents (local filesystem) ──────────────────────────────────────

@router.get("/courses/{course_id}/docs")
def list_docs(course_id: str, _user=Depends(verify_token)):
    rows = fetchall(
        "SELECT id, name, file_url, file_type, file_size, created_at, uploaded_by "
        "FROM course_document WHERE course_id = ? ORDER BY created_at DESC",
        (course_id,),
    )
    uploader_ids = list({r["uploaded_by"] for r in rows if r.get("uploaded_by")})
    name_map: dict = {}
    if uploader_ids:
        uph = ",".join("?" * len(uploader_ids))
        profs = fetchall(
            f"SELECT id, name, email FROM profiles WHERE id IN ({uph})", tuple(uploader_ids)
        )
        name_map = {p["id"]: p.get("name") or p.get("email", "") for p in profs}
    for r in rows:
        r["uploaded_by_name"] = name_map.get(r.get("uploaded_by", ""), "")
    return rows


@router.post("/courses/{course_id}/docs")
async def upload_doc(
    course_id: str,
    file: UploadFile = File(...),
    user=Depends(verify_token),
):
    require_coordinator_for_course(course_id, user)
    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 20 MB).")

    safe_filename = re.sub(r"[^\w.\-]", "_", file.filename or "file")
    rel_path      = f"{course_id}/{uuid.uuid4()}_{safe_filename}"
    full_path     = os.path.join(UPLOAD_DIR, rel_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    with open(full_path, "wb") as f:
        f.write(contents)
    file_url     = f"/uploads/{rel_path}"
    content_type = file.content_type or "application/octet-stream"

    doc_id = new_id()
    execute(
        "INSERT INTO course_document (id, course_id, name, file_url, file_type, file_size, uploaded_by) "
        "VALUES (?,?,?,?,?,?,?)",
        (doc_id, course_id, file.filename or safe_filename, file_url,
         content_type, len(contents), user.get("sub")),
    )
    return fetchone("SELECT * FROM course_document WHERE id = ?", (doc_id,))


@router.delete("/docs/{doc_id}")
def delete_doc(doc_id: str, user=Depends(verify_token)):
    row = fetchone("SELECT course_id, file_url FROM course_document WHERE id = ?", (doc_id,))
    if not row:
        raise HTTPException(404, "Document not found.")
    require_coordinator_for_course(row["course_id"], user)
    file_url: str = row.get("file_url", "")
    if file_url.startswith("/uploads/"):
        full_path = os.path.join(UPLOAD_DIR, file_url[len("/uploads/"):])
        try:
            if os.path.exists(full_path):
                os.remove(full_path)
        except OSError:
            pass
    execute("DELETE FROM course_document WHERE id = ?", (doc_id,))
    return {"message": "Document deleted."}


# ── Chatbot helper (used by chatbot.py) ───────────────────────────────────────

def get_projects_for_chatbot() -> list[dict]:
    """Return projects with nested course/semester/batch + students + guide profile for chatbot filter."""
    rows = fetchall("""
        SELECT p.project_id, p.title, p.github, p.guide, p.guide_id, p.readme_cache,
               pr.name AS guide_profile_name,
               c.course_name,
               s.sem_number,
               b.batch_name
        FROM project p
        LEFT JOIN profiles pr ON pr.id = p.guide_id
        LEFT JOIN course c ON c.course_id = p.course_id
        LEFT JOIN semester s ON s.semester_id = c.semester_id
        LEFT JOIN batch b ON b.batch_id = s.batch_id
        LIMIT 1000
    """)
    if not rows:
        return rows
    project_ids = [r["project_id"] for r in rows]
    ph = ",".join("?" * len(project_ids))
    students = fetchall(
        f"SELECT student_id, project_id, usn, name FROM student WHERE project_id IN ({ph})",
        tuple(project_ids),
    )
    stu_by_proj: dict = {}
    for s in students:
        stu_by_proj.setdefault(s["project_id"], []).append(
            {"student_id": s["student_id"], "usn": s["usn"], "name": s["name"]}
        )
    result = []
    for r in rows:
        result.append({
            "project_id":   r["project_id"],
            "title":        r["title"],
            "github":       r["github"],
            "guide":        r["guide"],
            "guide_id":     r["guide_id"],
            "readme_cache": r["readme_cache"],
            "profiles":     {"name": r["guide_profile_name"]} if r["guide_profile_name"] else {},
            "student":      stu_by_proj.get(r["project_id"], []),
            "course": {
                "course_name": r["course_name"],
                "semester": {
                    "sem_number": r["sem_number"],
                    "batch": {"batch_name": r["batch_name"]},
                },
            },
        })
    return result
